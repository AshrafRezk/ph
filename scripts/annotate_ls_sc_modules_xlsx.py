#!/usr/bin/env python3
"""Annotate LS & SC modules Excel with Life Sciences | Sales Core columns."""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path("/Users/ashrafrezk/Downloads/Modules Description (LS & SC).xlsx")
OUT = Path(
    "/Users/ashrafrezk/Salesforce Projects/Pharmaceuticals/Plan/"
    "Modules_Description_LS_SC_Annotated.xlsx"
)

# Life Sciences values
LS_NATIVE = "Native"
LS_ROADMAP = "Roadmap / Addon"
LS_CUSTOM = "Can be implemented"  # was Custom / Partner
LS_NOT = "Not in LS"

# Sales Core values
SC_DONE = "Native"  # was Already Implemented
SC_PARTIAL = "Partially Implemented"
SC_CAN = "Can be implemented"
SC_NA = "Not Applicable"

# (ls, sc, evidence) keyed by normalized "Module|Feature"
SCORES: dict[str, tuple[str, str, str]] = {
    # Sales
    "Sales|Sample Disbursement": (
        LS_NATIVE,
        SC_DONE,
        "Visit_Sample__c; Sample_Transaction__c (Distribute); visitCallShell",
    ),
    "Sales|Sample Limits": (
        LS_NATIVE,
        SC_CAN,
        "No sample-limit rules/objects; only on-hand qty check at distribute",
    ),
    "Sales|Cycle Plans": (
        LS_NATIVE,
        SC_PARTIAL,
        "planCycleManager + monthly account targets — not full product cycle-plan engine",
    ),
    "Sales|Sample Inventory": (
        LS_NATIVE,
        SC_DONE,
        "Sample_Inventory__c (Qty On Hand, Lot, Expiry, Owner)",
    ),
    "Sales|Field Coaching": (
        LS_NATIVE,
        SC_DONE,
        "Coaching_Template__c; Coaching_Event__c; Coaching_Module",
    ),
    "Sales|Lists and Filters": (
        LS_NATIVE,
        SC_DONE,
        "accountsTabOceList; Account_Rating_Layout__c; specialty/tier filters",
    ),
    "Sales|Sample Allocations": (
        LS_NATIVE,
        SC_CAN,
        "No HQ allocation object; Inventory Owner_User__c is stock ownership only",
    ),
    "Sales|Routing": (
        LS_NATIVE,
        SC_DONE,
        "fieldRepPlanner; plannerRouteUtils; Salesforce Maps",
    ),
    "Sales|Engagement Compliance (EPPV & MID)": (
        LS_NATIVE,
        SC_CAN,
        "No EPPV/MID compliance engine found",
    ),
    # Medical
    "Medical|Influence Graph Insight": (
        LS_NATIVE,
        SC_PARTIAL,
        "accountAffiliationNetwork — affiliation graph, not medical influence-insight AI",
    ),
    "Medical|Medical Education Plan Board": (
        LS_NATIVE,
        SC_CAN,
        "No MedEd plan board objects/LWCs in org",
    ),
    "Medical|Account & Product Level Planning": (
        LS_NATIVE,
        SC_PARTIAL,
        "Account_Territory_Product_Fields__c; Zeta_Project_Account_Goal__c — planning data, not MSL board",
    ),
    "Medical|Inquiry Triage Template": (
        LS_NATIVE,
        SC_PARTIAL,
        "Medical Inquiry via Case + visitMedicalInquiryModal — basic intake, limited triage templates",
    ),
    "Medical|Pre-Visit Planning": (
        LS_NATIVE,
        SC_PARTIAL,
        "Visit_Objective__c; fieldRepHomeTodayPlan; accountVisitInsightsPanel — prep views, not full MSL pack",
    ),
    "Medical|Insight Trending": (
        LS_NATIVE,
        SC_CAN,
        "No medical insight-trending module in org",
    ),
    "Medical|Voice-to-Insight": (
        LS_NATIVE,
        SC_CAN,
        "No voice/speech-to-insight feature; Agentforce would be extra license if added",
    ),
    "Medical|Scientific Asset Catalog": (
        LS_NATIVE,
        SC_PARTIAL,
        "CLM_Presentation__c content library covers detailing assets; not MSL scientific catalog",
    ),
    "Medical|Answer Framework": (
        LS_NATIVE,
        SC_CAN,
        "",
    ),
    "Medical|Next Best Action": (
        LS_NATIVE,
        SC_PARTIAL,
        "fieldRepHomeNextBestCustomer — Next Best Customer ranking, not full medical NBA",
    ),
    # KAM
    "KAM|HCO Profiling (Account 360)": (
        LS_NATIVE,
        SC_PARTIAL,
        "Institution RT + affiliations + visit history; not full KAM Account 360 workspace",
    ),
    "KAM|Account/Territory Plans/foundations": (
        LS_NATIVE,
        SC_PARTIAL,
        "Zeta_Project__c / Planning_Vision__c / targets — custom planning, not LSC Account Plan foundations",
    ),
    "KAM|Whole Office Management": (
        LS_NATIVE,
        SC_PARTIAL,
        "Account_Affiliation__c + visit attendees cover whole-office calls; limited KAM WOM tooling",
    ),
    "KAM|Account Plan Templates": (
        LS_NATIVE,
        SC_CAN,
        "No account-plan template library; Account_Plan__c is Quip URL only",
    ),
    "KAM|Account Team Roster": (
        LS_NATIVE,
        SC_CAN,
        "No Account Team roster module configured in Pharma apps",
    ),
    "KAM|Market Driven Engagement Plans": (
        LS_NATIVE,
        SC_CAN,
        "",
    ),
    "KAM|Engagement Prep": (
        LS_NATIVE,
        SC_CAN,
        "",
    ),
    "KAM|Annual Account Planning": (
        LS_NATIVE,
        SC_PARTIAL,
        "Zeta_Project__c annual/project planning exists; not dedicated annual account plan object",
    ),
    "KAM|Engagement Execution": (
        LS_NATIVE,
        SC_PARTIAL,
        "Zeta_Project_Activity__c + Visit__c tracking",
    ),
    "KAM|Project Management & Reporting": (
        LS_NATIVE,
        SC_DONE,
        "Zeta_Project__c suite; projectManagementHub; Zeta_C_Levels_App",
    ),
    # Marketing
    "Marketing|LSC Marketing App Experience": (
        LS_NATIVE,
        SC_CAN,
        "Pardot/Marketing Cloud packages present; no LSC Marketing app experience",
    ),
    "Marketing|Marketing Collaboration": (
        LS_NATIVE,
        SC_DONE,
        "Collaboration_Request__c; crossDeptCollaborationHub",
    ),
    "Marketing|Rep-Enrolled Campaigns": (
        LS_NATIVE,
        SC_CAN,
        "",
    ),
    "Marketing|Campaign Templates": (
        LS_NATIVE,
        SC_CAN,
        "Standard Campaign object available; no LS campaign template pack",
    ),
    # Common Capabilities
    "Common Capabilities|HCO & HCP Profiling": (
        LS_NATIVE,
        SC_DONE,
        "PersonAccount / Institution RTs; Pharma_Account_Module",
    ),
    "Common Capabilities|Addresses & Affiliations": (
        LS_NATIVE,
        SC_DONE,
        "Account_Affiliation__c; accountAffiliationNetwork; pharmacy affiliations noted",
    ),
    "Common Capabilities|Account Summarization": (
        LS_NATIVE,
        SC_CAN,
        "No AI account summarization; FieldPlanner AccountSummary is a map-pin DTO only",
    ),
    "Common Capabilities|Field Messaging": (
        LS_NATIVE,
        SC_DONE,
        "Home_Office_Message__c; homeOfficeMessages",
    ),
    "Common Capabilities|Authoring Territory Management - SPM": (
        LS_NATIVE,
        SC_PARTIAL,
        "territoryManagementConsole + Territory2 (ETM). Sales Planning (sfsp) is SDO demo package, not full SPM authoring",
    ),
    "Common Capabilities|Messages & Objectives": (
        LS_NATIVE,
        SC_DONE,
        "Visit_Objective__c; Visit_Product_Message__c; CLM_Message_Response__c",
    ),
    "Common Capabilities|Territory Alignment": (
        LS_NATIVE,
        SC_DONE,
        "Territory_Alignment_Module; Account_Territory_Fields__c",
    ),
    "Common Capabilities|Brick to Terr": (
        LS_NATIVE,
        SC_DONE,
        "Brick__c; Brick_Pharmacy__c; bricksManagementConsole",
    ),
    "Common Capabilities|Offline Reporting": (
        LS_NATIVE,
        SC_DONE,
        "Hybrid Briefcase + IndexedDB: Briefcase primes Visit/Account/Product/Coaching; IndexedDB caches CLM assets and queues Visit/Home/CLM/Planner/Coaching writes until sync",
    ),
    "Common Capabilities|Intelligent Content": (
        LS_NATIVE,
        SC_DONE,
        "CLM presentations / eDetailing (clmPlayer, CLM_* objects)",
    ),
    "Common Capabilities|Product Restrictions": (
        LS_NATIVE,
        SC_PARTIAL,
        "Product_Territory_Alignment__c scopes products; no full HCP-type product restriction rules",
    ),
    "Common Capabilities|Zip to Terr": (
        LS_NATIVE,
        SC_CAN,
        "Brick-to-territory exists; no dedicated zip/postal-to-territory mapper",
    ),
    "Common Capabilities|Affiliation Alignment": (
        LS_NATIVE,
        SC_DONE,
        "Account_Affiliation_Module; affiliation network UI",
    ),
    "Common Capabilities|Tableau Next Integration": (
        LS_NATIVE,
        SC_PARTIAL,
        "Tableau / Einstein shells present in org; no custom Tableau Next / CRMA datasets owned in Pharma modules",
    ),
    "Common Capabilities|OneKey Integration": (
        LS_CUSTOM,
        SC_PARTIAL,
        "adminConsole mentions OneKey connector; not a full OneKey sync module deployed",
    ),
    "Common Capabilities|Remote Engagement": (
        LS_NATIVE,
        SC_CAN,
        "No built-in video detailing / remote engagement tool",
    ),
    "Common Capabilities|Product Territory Alignment": (
        LS_NATIVE,
        SC_DONE,
        "Product_Territory_Alignment__c; productTerritoryManager",
    ),
    "Common Capabilities|Data Change Requests": (
        LS_NATIVE,
        SC_CAN,
        "No DCR workflow objects/UI in Pharma modules",
    ),
    "Common Capabilities|Consent & Preferences": (
        LS_NATIVE,
        SC_CAN,
        "Platform privacy stubs only; no HCP consent / preference center module",
    ),
    "Common Capabilities|Content API": (
        LS_NATIVE,
        SC_PARTIAL,
        "CLM uses ContentVersion downloads — not Life Sciences Content API product",
    ),
    "Common Capabilities|Segmentation & Targeting": (
        LS_NATIVE,
        SC_DONE,
        "Account_Rating_Layout__c; classification A/B/C; CLM territory targeting",
    ),
    "Common Capabilities|Product Catalog": (
        LS_NATIVE,
        SC_DONE,
        "Product2; Zeta_Product_Catalog; ChemipharmProductCatalogService",
    ),
    "Common Capabilities|Key Messages": (
        LS_NATIVE,
        SC_DONE,
        "Visit_Product_Message__c; CLM message feedback",
    ),
    "Common Capabilities|Time Off Territory": (
        LS_NATIVE,
        SC_DONE,
        "Time_Off_Request__c + TOT flows; Time_Off_Request_* permission sets",
    ),
    "Common Capabilities|Concur Integration": (
        LS_CUSTOM,
        SC_CAN,
        "No Concur connector configured",
    ),
    "Common Capabilities|Activity Scheduling": (
        LS_NATIVE,
        SC_DONE,
        "fieldRepPlanner; Visit__c scheduling; calendar planner",
    ),
    "Common Capabilities|Real-Time Collaboration": (
        LS_NATIVE,
        SC_PARTIAL,
        "Collaboration_Request__c is async cross-dept requests — not live co-editing / presence",
    ),
    "Common Capabilities|Activity Timeline": (
        LS_NATIVE,
        SC_DONE,
        "accountVisitInsightsPanel; CLM account activity history; standard timeline",
    ),
    "Common Capabilities|Intelligent Notifications": (
        LS_NATIVE,
        SC_PARTIAL,
        "Home office messages / KPI surfacing — not AI-triggered visit-gap notifications",
    ),
    "Common Capabilities|Briefings (Stories)": (
        LS_NATIVE,
        SC_CAN,
        "No Briefings/Stories feature; Agentforce would be extra license if added",
    ),
    # Order Management
    "Order Management|Store Check": (LS_NATIVE, SC_CAN, ""),
    "Order Management|Free Goods": (
        LS_NATIVE,
        SC_CAN,
        "No free-goods module in Pharma field apps",
    ),
    "Order Management|Pricing": (
        LS_NATIVE,
        SC_PARTIAL,
        "Standard price books / catalog seeding — no field order pricing engine",
    ),
    "Order Management|Split Deliveries": (LS_NATIVE, SC_CAN, ""),
    "Order Management|Direct Orders": (
        LS_NATIVE,
        SC_CAN,
        "SDO OMS console may exist; no Pharma field direct-order UX",
    ),
    "Order Management|Multi-Recipient Orders": (LS_NATIVE, SC_CAN, ""),
    "Order Management|Transfer Orders": (LS_NATIVE, SC_CAN, ""),
    "Order Management|Discounts": (
        LS_NATIVE,
        SC_CAN,
        "No discount module in Pharma field apps",
    ),
    "Order Management|Contracts": (LS_NATIVE, SC_CAN, ""),
    "Order Management|Promotions": (
        LS_NATIVE,
        SC_PARTIAL,
        "Promo_Budget__c / Promo_Budget_Line__c — promo budgets, not full order promotions engine",
    ),
    # Events
    "Events|Attendee Recruitment": (LS_ROADMAP, SC_CAN, ""),
    "Events|Budget Management": (
        LS_ROADMAP,
        SC_PARTIAL,
        "Promo_Budget__c / Zeta_Project_Budget_Line__c — not events-specific budget module",
    ),
    "Events|Monetary Cap & Utilization Limits": (LS_ROADMAP, SC_CAN, ""),
    "Events|Expert Engagement": (
        LS_ROADMAP,
        SC_CAN,
        "No expert engagement / speaker management module in Pharma apps",
    ),
    "Events|Expert Onboarding": (
        LS_ROADMAP,
        SC_CAN,
        "No expert onboarding module in Pharma apps",
    ),
    "Events|Country Regulations & Practices": (LS_ROADMAP, SC_CAN, ""),
    "Events|Expense Management": (
        LS_ROADMAP,
        SC_CAN,
        "No expense management module; Concur also not configured",
    ),
    "Events|Agency Services": (LS_ROADMAP, SC_CAN, ""),
    "Events|Fee for Service & Payment Requests": (LS_ROADMAP, SC_CAN, ""),
    "Events|Logistics Needs": (LS_ROADMAP, SC_CAN, ""),
    "Events|Flow of Work Compliance": (LS_ROADMAP, SC_CAN, ""),
    "Events|Governing Agreements": (LS_ROADMAP, SC_CAN, ""),
    "Events|AOP & Needs Assessments": (LS_ROADMAP, SC_CAN, ""),
    "Events|FMV & ToV": (LS_ROADMAP, SC_CAN, ""),
    "Events|Document Templates": (LS_ROADMAP, SC_CAN, ""),
}

# Extra modules not in source Excel (Module, Feature, Description, ls, sc, evidence)
EXTRA_ROWS: list[tuple[str, str, str, str, str, str]] = [
    (
        "Sales",
        "Call / Visit Reporting",
        "Log HCP/HCO visits with attendees, products discussed, samples, notes, and submit status. Offline: visit must be cached while online first (or created offline via planner queue); then call report can queue via IndexedDB until sync.",
        LS_NATIVE,
        SC_DONE,
        "Visit__c; visitCallShell; VisitCallReportService; offline queue in clmOfflineStore",
    ),
    (
        "Sales",
        "Visit Attendees",
        "Capture multiple HCPs / office staff who participated in a call.",
        LS_NATIVE,
        SC_DONE,
        "Visit_Attendee__c",
    ),
    (
        "Sales",
        "Product Detailing",
        "Record which products were detailed on the call with related messages.",
        LS_NATIVE,
        SC_DONE,
        "Visit_Product_Detail__c; Visit_Product_Message__c",
    ),
    (
        "Sales",
        "CLM / eDetailing",
        "Present approved digital detailing content and capture slide/message metrics. Offline uses IndexedDB asset cache (Briefcase for CRM records): prefetch presentations while online, then play/capture offline until sync.",
        LS_NATIVE,
        SC_DONE,
        "CLM_* objects; clmPlayer; CLM_Module; clmOfflineStore IndexedDB; Pharma_Field_Rep_Offline Briefcase",
    ),
    (
        "Sales",
        "Call Surveys / Feedback",
        "Capture structured survey responses or ratings during/after a visit.",
        LS_NATIVE,
        SC_DONE,
        "Visit_Product_Survey_Feedback__c; Visit_Survey_Module",
    ),
    (
        "Sales",
        "Next Best Customer",
        "Rank accounts to prioritize next visits based on coverage and potential.",
        LS_NATIVE,
        SC_DONE,
        "fieldRepHomeNextBestCustomer",
    ),
    (
        "Sales",
        "Sample Signature / Acknowledgement",
        "Capture HCP acknowledgement/signature when samples are disbursed (compliance).",
        LS_NATIVE,
        SC_CAN,
        "Disbursement exists; no signature capture object/UI found",
    ),
    (
        "Sales",
        "Inventory Transfer & Adjustment",
        "Move or adjust rep sample stock (transfer between reps, spoilage, return to warehouse).",
        LS_NATIVE,
        SC_PARTIAL,
        "Sample_Transaction__c types: Distribute, Return, Adjustment — transfer between reps limited",
    ),
    (
        "Sales",
        "Field Rep Home Dashboard",
        "Rep landing page with KPIs, today's plan/map, and priorities. Today plan and Home metrics/NBC can show cached views offline after online prefetch.",
        LS_NATIVE,
        SC_DONE,
        "fieldRepHomeMetrics; fieldRepHomeTodayPlan; Field_Rep_Home",
    ),
    (
        "Sales",
        "Medical Rep 360 / Manager KPI",
        "Manager view of rep coverage, call quality, coaching scores, and CLM usage.",
        LS_NATIVE,
        SC_DONE,
        "Medical_Rep_360_Dashboard; managementTeamKpiDashboard; coaching insights",
    ),
    (
        "Medical",
        "Medical Inquiry Intake",
        "Capture HCP medical questions from the field and route to Medical Affairs.",
        LS_NATIVE,
        SC_DONE,
        "visitMedicalInquiryModal; Case Medical Inquiry RT; Medical_Inquiry_Module",
    ),
    (
        "Medical",
        "Adverse Event Capture / EPPV Handoff",
        "Flag and hand off potential adverse event mentions from field interactions to PV.",
        LS_NATIVE,
        SC_CAN,
        "Related to Engagement Compliance (EPPV); not implemented as dedicated AE module",
    ),
    (
        "Common Capabilities",
        "Offline Mobile Sync (CLM)",
        "Hybrid Briefcase + IndexedDB. Briefcase syncs Visit/Account/Product/Coaching records for mobile; IndexedDB prefetches CLM assets and queues call-report, planner upsert/reschedule, and coaching create/submit until sync.",
        LS_NATIVE,
        SC_DONE,
        "Pharma_Field_Rep_Offline Briefcase; clmOfflineStore; fieldRepHomeClmPrefetch; ClmOfflineSyncController",
    ),
    (
        "Common Capabilities",
        "Approved / Compliant Email",
        "Send pre-approved content to HCPs with consent and tracking.",
        LS_NATIVE,
        SC_CAN,
        "",
    ),
    (
        "Common Capabilities",
        "Coverage & Frequency Analytics",
        "Track achieved vs target visits by account classification (A/B/C) and RF metrics.",
        LS_NATIVE,
        SC_DONE,
        "fieldRepHomeMetrics coverage/RF%; Employee_Time_Card targets",
    ),
    (
        "Common Capabilities",
        "Rep GPS / Fleet Tracking",
        "Capture rep location snapshots and management fleet map view.",
        LS_CUSTOM,
        SC_DONE,
        "Rep_Location_Snapshot__c; managementFleetMap; repLocationPublisher",
    ),
    (
        "Analytics",
        "Pharmacy Sales Analytics",
        "Import pharmacy sales/withdrawal data and surface ROI recommendations to the field.",
        LS_NOT,
        SC_DONE,
        "Pharmacy_Sales_* objects; pharmacySalesDashboard; Pharmacy_Sales_Analytics",
    ),
    (
        "Analytics",
        "Promo Budget Tracking",
        "Plan and track promotional spend / budget lines at HQ or project level.",
        LS_ROADMAP,
        SC_DONE,
        "Promo_Budget__c; Promo_Budget_Line__c; promoBudgetDashboard",
    ),
    (
        "KAM",
        "Account Objectives & Tactics",
        "Break account plans into measurable objectives and executable tactics.",
        LS_NATIVE,
        SC_PARTIAL,
        "Zeta_Project_KPI__c / Account Goal — custom project goals, not LSC AccountPlanObjective",
    ),
    (
        "Common Capabilities",
        "External Integration Hub (Mendix / IMS)",
        "Monitor and manage connectors to Mendix, IMS Health, Maps, and similar platforms.",
        LS_CUSTOM,
        SC_PARTIAL,
        "Mendix_Sync_Log__c; mendixIntegrationHub; integrationsManagementConsole",
    ),
    (
        "Analytics",
        "CRM Analytics (CRMA)",
        "Salesforce CRM Analytics dashboards and datasets for field and commercial insights.",
        LS_NATIVE,
        SC_PARTIAL,
        "Analytics shells may exist in org; no custom Pharma CRMA datasets/dashboards in project modules",
    ),
    (
        "Sales",
        "S&OP",
        "Sales & Operations Planning — align demand, forecast, and supply targets across commercial and ops.",
        LS_NATIVE,
        SC_CAN,
        "No S&OP module in Pharma apps",
    ),
    (
        "Order Management",
        "Distribution Management App",
        "Back-office distribution management for distributors, inventory, and commercial ops.",
        LS_CUSTOM,
        SC_CAN,
        "Separate F&B Distribution plan exists as design work — not live in Pharma Sales Core",
    ),
    (
        "Order Management",
        "Distribution Field Sales App",
        "Field sales application for distribution reps (visits, orders, stock, and route execution).",
        LS_CUSTOM,
        SC_CAN,
        "Separate F&B Distribution plan exists as design work — not live in Pharma Sales Core",
    ),
]


DESC_OVERRIDES: dict[str, str] = {
    "Common Capabilities|Offline Reporting": (
        "Offline is Implemented via hybrid Salesforce Briefcase (mobile CRM record sync) plus custom IndexedDB "
        "(CLM assets and write-ahead action queue). After online prefetch/cache, Visit reporting, Home (Today Plan + "
        "metrics), CLM, Planner create/reschedule, and Coaching create/submit work offline until reconnect and sync."
    ),
    "Sales|Call / Visit Reporting": (
        "Log HCP/HCO visits with attendees, products discussed, samples, notes, and submit status. "
        "Offline: visit must be cached while online first (or created offline via planner queue); then call report can queue via IndexedDB until sync."
    ),
}


def key(module: str | None, feature: str | None) -> str:
    return f"{(module or '').strip()}|{(feature or '').strip()}"


def score_for(module: str, feature: str) -> tuple[str, str, str]:
    k = key(module, feature)
    if k in SCORES:
        return SCORES[k]
    # fallback: treat catalog CE feature as Native / Can Be Implemented
    return (LS_NATIVE, SC_CAN, "")


FILL_SC = {
    SC_DONE: PatternFill("solid", fgColor="C6EFCE"),
    SC_PARTIAL: PatternFill("solid", fgColor="FFEB9C"),
    SC_CAN: PatternFill("solid", fgColor="DDEBF7"),
    SC_NA: PatternFill("solid", fgColor="E7E6E6"),
}
FILL_LS = {
    LS_NATIVE: PatternFill("solid", fgColor="E2EFDA"),
    LS_ROADMAP: PatternFill("solid", fgColor="FCE4D6"),
    LS_CUSTOM: PatternFill("solid", fgColor="DDEBF7"),
    LS_NOT: PatternFill("solid", fgColor="F2F2F2"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def style_header(ws, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(12, length + 2)


def main() -> None:
    src = load_workbook(SRC)
    src_ws = src.active

    wb = Workbook()
    ws = wb.active
    ws.title = "Modules"

    headers = [
        "Module",
        "Feature",
        "Zeta Description",
        "Selected",
        "Life Sciences",
        "Sales Core",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"

    out_row = 2
    seen: set[str] = set()

    for row in src_ws.iter_rows(min_row=2, values_only=True):
        module = (row[0] or "").strip() if row[0] else ""
        feature = (row[1] or "").strip() if row[1] else ""
        desc = row[2] if len(row) > 2 else None
        selected = row[3] if len(row) > 3 else None
        if not module and not feature:
            continue
        if module == "Module" or feature == "Feature":
            continue

        ls, sc, _evidence = score_for(module, feature)
        seen.add(key(module, feature))
        desc = DESC_OVERRIDES.get(key(module, feature), desc)

        values = [module, feature, desc, selected, ls, sc]
        for col, val in enumerate(values, 1):
            cell = ws.cell(out_row, col, val)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(out_row, 5).fill = FILL_LS.get(ls, PatternFill())
        ws.cell(out_row, 6).fill = FILL_SC.get(sc, PatternFill())
        out_row += 1

    # Append extras not already present
    for module, feature, desc, ls, sc, _evidence in EXTRA_ROWS:
        k = key(module, feature)
        if k in seen:
            continue

        values = [module, feature, desc, None, ls, sc]
        for col, val in enumerate(values, 1):
            cell = ws.cell(out_row, col, val)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(out_row, 5).fill = FILL_LS.get(ls, PatternFill())
        ws.cell(out_row, 6).fill = FILL_SC.get(sc, PatternFill())
        # Tag newly added rows lightly on Module cell
        ws.cell(out_row, 1).font = Font(italic=True)
        out_row += 1
        seen.add(k)

    autosize(ws)
    ws.column_dimensions["C"].width = 55
    ws.row_dimensions[1].height = 30

    counts_sc: dict[str, int] = {}
    counts_ls: dict[str, int] = {}
    for r in range(2, out_row):
        sc = ws.cell(r, 6).value
        ls = ws.cell(r, 5).value
        counts_sc[sc] = counts_sc.get(sc, 0) + 1
        counts_ls[ls] = counts_ls.get(ls, 0) + 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Rows: {out_row - 2}")
    print("Sales Core:", counts_sc)
    print("Life Sciences:", counts_ls)

    # Validate all source features were scored explicitly
    missing_explicit = []
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        module = (row[0] or "").strip() if row[0] else ""
        feature = (row[1] or "").strip() if row[1] else ""
        if not feature or feature == "Feature":
            continue
        if key(module, feature) not in SCORES:
            missing_explicit.append(key(module, feature))
    if missing_explicit:
        print("WARNING: source features without explicit score (used fallback):")
        for m in missing_explicit:
            print(" ", m)


if __name__ == "__main__":
    main()
